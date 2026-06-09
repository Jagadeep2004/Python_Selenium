import openpyxl

def get_data(path, sheet_name):
    final_list = []

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(2, rows + 1):
        row = []

        for c in range(1, cols + 1):
            row.append(sheet.cell(r, c).value)

        final_list.append(row)

    return final_list